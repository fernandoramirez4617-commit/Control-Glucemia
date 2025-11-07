from flask import Flask, request, jsonify, Response, send_file, send_from_directory
from flask_cors import CORS
import sqlite3, csv, io, os, math, traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# ==========================
# Config compatible Windows/Render
# ==========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TMP_DIR = "/tmp" if os.name != "nt" else BASE_DIR  # En Windows guardamos en el propio backend

DB_PATH   = os.getenv("DB_PATH", os.path.join(BASE_DIR, "data.db"))
XLSX_PATH = os.getenv("XLSX_PATH", os.path.join(TMP_DIR, "patients.xlsx"))
PDF_PATH  = os.getenv("PDF_PATH",  os.path.join(TMP_DIR, "patients.pdf"))

FRONTEND_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", "frontend"))
app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path="")
CORS(app)
app.url_map.strict_slashes = False  # ✅ Acepta rutas con y sin slash final

# ==========================
# DB helpers
# ==========================
def get_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True) if os.path.dirname(DB_PATH) else None
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS patients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            age INTEGER,
            sex TEXT,
            schooling TEXT,
            glucose_mgdl REAL,
            risk TEXT,
            has_hypertension INTEGER,
            has_obesity INTEGER,
            has_dyslipidemia INTEGER,
            has_ckd INTEGER,
            has_cvd INTEGER,
            has_copd_asthma INTEGER,
            has_depression INTEGER,
            systolic INTEGER,
            diastolic INTEGER,
            htn_stage TEXT,
            weight_kg REAL,
            height_cm REAL,
            bmi REAL,
            bmi_cat TEXT,
            smoker INTEGER,
            physical_activity TEXT,
            med_htn INTEGER,
            med_dm INTEGER,
            med_insulin INTEGER,
            med_metformin INTEGER,
            med_statins INTEGER,
            med_antiplatelet INTEGER,
            med_other TEXT,
            notes TEXT,
            created_at TEXT
        )
        """
    )
    conn.commit()
    conn.close()
    export_xlsx()

# ==========================
# Funciones de cálculo
# ==========================
def compute_risk(glucose):
    try: g = float(glucose)
    except (TypeError, ValueError): return "desconocido"
    if g < 70: return "alto (hipoglucemia)"
    if 70 <= g <= 99: return "bajo"
    if 100 <= g <= 125: return "moderado"
    if g >= 126: return "alto"
    return "desconocido"

def compute_bmi(weight_kg, height_cm):
    try:
        w = float(weight_kg) if weight_kg else None
        h = float(height_cm) / 100.0 if height_cm else None
    except (TypeError, ValueError):
        return None, None
    if not w or not h or h == 0: return None, None
    bmi = w / (h*h)
    if bmi < 18.5: cat = "bajo peso"
    elif bmi < 25: cat = "normal"
    elif bmi < 30: cat = "sobrepeso"
    else: cat = "obesidad"
    return round(bmi, 1), cat

def compute_htn_stage(sys, dia):
    try:
        s = int(sys) if sys else None
        d = int(dia) if dia else None
    except (TypeError, ValueError):
        return None
    if s is None or d is None: return None
    if s < 120 and d < 80: return "normal"
    if 120 <= s <= 129 and d < 80: return "elevada"
    if (130 <= s <= 139) or (80 <= d <= 89): return "HTA grado 1"
    if s >= 140 or d >= 90: return "HTA grado 2"
    return None

# ==========================
# Query helpers
# ==========================
def _filters_to_where(filters):
    where, params = [], []
    if filters.get("risk"): where.append("LOWER(risk)=LOWER(?)"); params.append(filters["risk"])
    if filters.get("name"): where.append("LOWER(name) LIKE LOWER(?)"); params.append(f"%{filters['name']}%")
    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    return where_sql, params

def rows_to_list(filters=None, limit=None, offset=None):
    filters = filters or {}
    where_sql, params = _filters_to_where(filters)
    limit_sql = ""
    if limit and offset is not None:
        limit_sql = " LIMIT ? OFFSET ?"; params += [limit, offset]
    conn = get_db()
    rows = conn.execute("SELECT * FROM patients"+where_sql+" ORDER BY id DESC"+limit_sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def count_rows(filters=None):
    filters = filters or {}
    where_sql, params = _filters_to_where(filters)
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) c FROM patients"+where_sql, params).fetchone()["c"]
    conn.close()
    return n

# ==========================
# Exportaciones
# ==========================
def export_xlsx():
    data = rows_to_list()
    wb = Workbook(); ws = wb.active; ws.title = "patients"
    headers = [col for col in data[0].keys()] if data else []
    if not headers:
        headers = ["id","name","age","sex","schooling","glucose_mgdl","risk"]
    ws.append(headers)
    for row in data: ws.append([row.get(h) for h in headers])
    for i in range(1, len(headers)+1): ws.column_dimensions[get_column_letter(i)].width = 14
    wb.save(XLSX_PATH)

def export_pdf():
    data = rows_to_list()
    pdf = SimpleDocTemplate(PDF_PATH, pagesize=landscape(A4), title="Pacientes")
    styles = getSampleStyleSheet()
    if not data:
        pdf.build([Paragraph("No hay registros para exportar.", styles["Title"])]); return
    headers = ["ID","Fecha","Nombre","Edad","Sexo","Escolaridad","Glucemia","Riesgo"]
    rows = []
    for r in data:
        rows.append([r.get("id"), (r.get("created_at") or "")[:19], r.get("name",""),
                     r.get("age",""), r.get("sex",""), r.get("schooling",""),
                     r.get("glucose_mgdl",""), r.get("risk","")])
    table = Table([headers] + rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#004C70")),
        ("TEXTCOLOR",(0,0),(-1,0), colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,0),10),
        ("GRID",(0,0),(-1,-1),0.25, colors.grey)
    ]))
    pdf.build([table])

# ==========================
# Utilidades
# ==========================
def _b(v): return 1 if v else 0

# ==========================
# Rutas API
# ==========================
@app.route("/api/health")
def health(): return jsonify({"status":"ok","time":datetime.utcnow().isoformat()+"Z"})

@app.route("/api/patients", methods=["GET", "POST"])
@app.route("/api/patients/", methods=["GET", "POST"])
def patients_handler():
    if request.method == "GET":
        risk = request.args.get("risk"); name = request.args.get("name")
        page = int(request.args.get("page", 1)); page_size = int(request.args.get("page_size", 10))
        page = max(1, page); page_size = max(1, min(page_size, 100))
        filters = {"risk": risk, "name": name}
        total = count_rows(filters)
        offset = (page-1)*page_size
        items = rows_to_list(filters, limit=page_size, offset=offset)
        return jsonify({"items": items, "total": total, "page": page, "page_size": page_size, "pages": math.ceil(total/page_size)})

    # POST
    d = request.get_json(force=True)
    name = (d.get("name") or "").strip()
    if not name: return jsonify({"error":"El nombre es obligatorio."}), 400
    try: g = float(d.get("glucose_mgdl"))
    except (TypeError, ValueError): return jsonify({"error":"La glucemia (mg/dL) debe ser numérica."}), 400

    risk = compute_risk(g)
    bmi, bmi_cat = compute_bmi(d.get("weight_kg"), d.get("height_cm"))
    htn_stage = compute_htn_stage(d.get("systolic"), d.get("diastolic"))
    created_at = datetime.utcnow().isoformat() + "Z"

    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO patients (name, age, sex, schooling, glucose_mgdl, risk,
        has_hypertension, has_obesity, has_dyslipidemia, has_ckd, has_cvd,
        has_copd_asthma, has_depression, systolic, diastolic, htn_stage,
        weight_kg, height_cm, bmi, bmi_cat, smoker, physical_activity,
        med_htn, med_dm, med_insulin, med_metformin, med_statins,
        med_antiplatelet, med_other, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [d.get("name"), d.get("age"), d.get("sex"), d.get("schooling"), g, risk,
          _b(d.get("has_hypertension")), _b(d.get("has_obesity")), _b(d.get("has_dyslipidemia")), _b(d.get("has_ckd")),
          _b(d.get("has_cvd")), _b(d.get("has_copd_asthma")), _b(d.get("has_depression")),
          d.get("systolic"), d.get("diastolic"), htn_stage, d.get("weight_kg"), d.get("height_cm"),
          bmi, bmi_cat, _b(d.get("smoker")), d.get("physical_activity"),
          _b(d.get("med_htn")), _b(d.get("med_dm")), _b(d.get("med_insulin")), _b(d.get("med_metformin")),
          _b(d.get("med_statins")), _b(d.get("med_antiplatelet")), d.get("med_other"), d.get("notes"), created_at])
    conn.commit(); conn.close()
    export_xlsx(); export_pdf()
    return jsonify({"ok": True})

@app.route("/api/patients/<int:pid>", methods=["PUT", "DELETE"])
def modify_patient(pid):
    conn = get_db()
    if request.method == "DELETE":
        conn.execute("DELETE FROM patients WHERE id=?", (pid,))
        conn.commit(); conn.close()
        export_xlsx(); export_pdf()
        return jsonify({"ok": True})

    # PUT
    d = request.get_json(force=True)
    to_set = {}
    for k,v in d.items():
        if k.startswith("has_") or k.startswith("med_") or k in ["smoker"]:
            to_set[k] = 1 if v else 0
        else:
            to_set[k] = v
    if "glucose_mgdl" in d:
        try: g=float(d["glucose_mgdl"])
        except (TypeError, ValueError): return jsonify({"error":"La glucemia debe ser numérica."}), 400
        to_set["risk"] = compute_risk(g)
    cols=", ".join([f"{k}=?" for k in to_set.keys()]); vals=list(to_set.values())+[pid]
    conn.execute(f"UPDATE patients SET {cols} WHERE id=?", vals)
    conn.commit(); conn.close()
    export_xlsx(); export_pdf()
    return jsonify({"ok": True})

# ==========================
# Estadísticas
# ==========================
@app.route("/api/stats", methods=["GET"])
@app.route("/api/stats/", methods=["GET"])
def stats():
    conn = get_db()
    cur = conn.cursor()
    total = cur.execute("SELECT COUNT(*) c FROM patients").fetchone()["c"]
    by_risk = {r["risk"] or "desconocido": r["c"] for r in cur.execute("SELECT risk, COUNT(*) c FROM patients GROUP BY risk").fetchall()}
    by_bmi = {r["bmi_cat"] or "sin dato": r["c"] for r in cur.execute("SELECT bmi_cat, COUNT(*) c FROM patients GROUP BY bmi_cat").fetchall()}
    with_htn = cur.execute("SELECT COUNT(*) c FROM patients WHERE has_hypertension=1").fetchone()["c"]
    with_obesity = cur.execute("SELECT COUNT(*) c FROM patients WHERE has_obesity=1").fetchone()["c"]
    conn.close()
    return jsonify({"total": total, "by_risk": by_risk, "by_bmi": by_bmi, "with_hypertension": with_htn, "with_obesity": with_obesity})

# ==========================
# Frontend
# ==========================
@app.route("/")
def root():
    return send_from_directory(app.static_folder, "index.html")

# ==========================
# Error handler
# ==========================
@app.errorhandler(Exception)
def handle_exception(e):
    print("ERROR:", repr(e))
    traceback.print_exc()
    return jsonify({"error": "server_error", "detail": str(e)}), 500

# ==========================
# Init & Run
# ==========================
with app.app_context():
    init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
